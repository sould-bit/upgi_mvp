from datetime import date, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.domains.auth.utils import get_current_admin
from app.domains.users.models import User
from app.domains.reportes.service import ReporteService
from app.domains.reportes.schemas import (
    DashboardResponse, ReporteSemanaResponse, ReporteIngresosResponse,
    AdminReservaListResponse, OcupacionResponse, HorariosPicoResponse,
    ClientesFrecuentesResponse, DailyResponse
)
from app.domains.reservas.service import ReservaService

router = APIRouter(prefix="/admin", tags=["Admin"])


@router.get("/dashboard", response_model=DashboardResponse)
def get_dashboard(current_user: User = Depends(get_current_admin)):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        service = ReporteService(db)
        return service.get_stats()
    finally:
        db.close()


@router.get("/reportes/reservas-semana", response_model=ReporteSemanaResponse)
def get_reporte_semana(
    fecha_inicio: date = Query(...),
    fecha_fin: date = Query(...),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        service = ReporteService(db)
        return service.get_reservas_semana(fecha_inicio, fecha_fin)
    finally:
        db.close()


@router.get("/reportes/ingresos", response_model=ReporteIngresosResponse)
def get_reporte_ingresos(
    fecha_desde: date = Query(...),
    fecha_hasta: date = Query(...),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        service = ReporteService(db)
        return service.get_ingresos(fecha_desde, fecha_hasta)
    finally:
        db.close()


@router.get("/reservas", response_model=AdminReservaListResponse)
def listar_todas_reservas(
    fecha: date | None = Query(None),
    cancha_id: int | None = Query(None),
    estado_pago: str | None = Query(None),
    usuario_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        service = ReservaService(db)
        return service.listar_todas(
            fecha=fecha,
            cancha_id=cancha_id,
            estado_pago=estado_pago,
            usuario_id=usuario_id,
            page=page,
            limit=limit
        )
    finally:
        db.close()


@router.get("/reportes/ocupacion", response_model=OcupacionResponse)
def get_reporte_ocupacion(
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    cancha_id: int | None = Query(default=None),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        if fecha_desde is None:
            fecha_desde = date.today() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = date.today()
        if fecha_desde > fecha_hasta:
            raise HTTPException(status_code=400, detail="fecha_desde must be <= fecha_hasta")

        service = ReporteService(db)
        return service.get_ocupacion(fecha_desde, fecha_hasta, cancha_id)
    finally:
        db.close()


@router.get("/reportes/horarios-pico", response_model=HorariosPicoResponse)
def get_reporte_horarios_pico(
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    cancha_id: int | None = Query(default=None),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        if fecha_desde is None:
            fecha_desde = date.today() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = date.today()
        if fecha_desde > fecha_hasta:
            raise HTTPException(status_code=400, detail="fecha_desde must be <= fecha_hasta")

        service = ReporteService(db)
        return service.get_horarios_pico(fecha_desde, fecha_hasta, cancha_id)
    finally:
        db.close()


@router.get("/reportes/clientes-frecuentes", response_model=ClientesFrecuentesResponse)
def get_reporte_clientes(
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    cancha_id: int | None = Query(default=None),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        if fecha_desde is None:
            fecha_desde = date.today() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = date.today()
        if fecha_desde > fecha_hasta:
            raise HTTPException(status_code=400, detail="fecha_desde must be <= fecha_hasta")

        service = ReporteService(db)
        return service.get_clientes_frecuentes(fecha_desde, fecha_hasta, cancha_id)
    finally:
        db.close()


@router.get("/reportes/daily", response_model=DailyResponse)
def get_reporte_daily(
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    cancha_id: int | None = Query(default=None),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        if fecha_desde is None:
            fecha_desde = date.today() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = date.today()
        if fecha_desde > fecha_hasta:
            raise HTTPException(status_code=400, detail="fecha_desde must be <= fecha_hasta")

        service = ReporteService(db)
        return service.get_daily(fecha_desde, fecha_hasta, cancha_id)
    finally:
        db.close()


@router.get("/reportes/export/excel")
def exportar_reportes_excel(
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    cancha_id: int | None = Query(default=None),
    current_user: User = Depends(get_current_admin)
):
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        if fecha_desde is None:
            fecha_desde = date.today() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = date.today()
        if fecha_desde > fecha_hasta:
            raise HTTPException(status_code=400, detail="fecha_desde must be <= fecha_hasta")

        service = ReporteService(db)

        ocupacion_data = service.get_ocupacion(fecha_desde, fecha_hasta, cancha_id)
        horarios_data = service.get_horarios_pico(fecha_desde, fecha_hasta, cancha_id)
        clientes_data = service.get_clientes_frecuentes(fecha_desde, fecha_hasta, cancha_id)
        daily_data = service.get_daily(fecha_desde, fecha_hasta, cancha_id)

        import openpyxl

        workbook = openpyxl.Workbook()

        ws_daily = workbook.active
        ws_daily.title = "Daily"
        ws_daily.append(["Fecha", "Reservas", "Ingreso Total"])
        for item in daily_data.get("daily", []):
            ws_daily.append([item["fecha"], item["reservas_count"], item["ingreso_total"]])

        ws_ocup = workbook.create_sheet("Ocupacion")
        ws_ocup.append(["Cancha", "Horas Reservadas", "Horas Disponibles", "Ocupación %"])
        for item in ocupacion_data.get("ocupacion", []):
            ws_ocup.append(
                [
                    item["cancha_nombre"],
                    item["horas_reservadas"],
                    item["horas_disponibles"],
                    item["ocupacion_pct"],
                ]
            )

        ws_horarios = workbook.create_sheet("HorariosPico")
        ws_horarios.append(["Hora", "Cantidad de Reservas"])
        for item in horarios_data.get("horarios", []):
            ws_horarios.append([item["hora"], item["cantidad"]])

        ws_clientes = workbook.create_sheet("ClientesFrecuentes")
        ws_clientes.append(["Cliente", "Total Reservas", "Total Gastado"])
        for item in clientes_data.get("clientes", []):
            ws_clientes.append([item["cliente_nombre"], item["total_reservas"], item["total_gastado"]])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"reportes_{fecha_desde}_{fecha_hasta}.xlsx"
        headers = {"Content-Disposition": f"attachment; filename={filename}"}

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    finally:
        db.close()
